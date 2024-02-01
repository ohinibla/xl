import asyncio
import random
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from rich import print
from rich.progress import track


def generate_random_num() -> str:
    """
    Generate a random phone number.

    Returns:
       - str: A string representing a random phone number with known prefixes.
    """

    # typical Iran's cellphone number prefix
    prefix = random.choice(["12", "35", "36", "02", "18"])
    num = [str(i) for i in range(7)]
    random.shuffle(num)
    num = "".join(num)
    num = f"09{prefix}{num}"
    return num


def generate_days(days: int, numbers: int) -> None:
    """
    Generate days[int] with numbers[int] amount of random numbers.

    Input:
       - days: int
       - numbers: int

    Returns:
       - None -> saves the output in .xlsx files in "months" folder (create if not found) in the current directory

    """

    (Path(".") / "months").mkdir(exist_ok=True)
    for i in track(range(1, days + 1), description="generating test cases         "):
        wb = Workbook()
        ws = wb.active
        for row in ws.iter_rows(min_row=1, max_col=1, max_row=numbers):
            for cell in row:
                cell.value = generate_random_num()
        wb.save(Path(".") / "months" / f"{i}.xlsx")


def get_files_values(files: set) -> dict[str, tuple[(str, list)]]:
    """
    Collect values of row A of every files and create a dictionary based on their frequency

    Input:
       - files: set -> set of .xlsx file paths

    Returns:
       - dict[str: tuple] -> a dictionary of str: tuple with numbers (as string) and their
       (frequency, a list of files that their are contained) tuple
    """

    # values dictionary will be populated with tuples of int, list of number and their location, i.e: {(1, [1.xlsx, 2.xlsx])}
    values = {}
    for xp in track(files, description="generating values dictionary  "):
        wb = load_workbook(filename=xp, data_only=True)
        ws = wb.active
        col_A = ws["A"]
        for v in col_A:
            _v = v.value
            # if not found return a a tuple of (0, [])
            num = values.get(_v, (0, []))[0] + 1
            fn = values.get(_v, (0, []))[1] + [str(xp)]
            values[_v] = (num, fn)

    return values


async def edit_files_values(
    valuesDict: dict,
    files: set,
    make_copy: bool,
    show_dup_origin: bool,
    test: bool,
) -> None:
    """
    Asynchronously generate new files or rewrite existing with the duplicated numbers marked
    with a different style and their occurrences in the cell to their right
    (used async so that the progressbar can be update in the gui with every file write)

    Input:
        - valuesDict: dictionary
        files: set
        make_copy: bool
        show_dup_origin: bool
        test: bool

    Returns:
        - None -> saved new or existing files with duplicate numbers marked
    """

    # TODO: use existing Path methods to handle paths
    path_delimiter = "\\" if test else "/"
    fp = Path(".") / "data"
    for xp in track(files, description="writing files                 "):
        wb = load_workbook(filename=xp, data_only=True)
        ws = wb.active
        col_A = ws["A"]
        for cell in col_A:
            if valuesDict[cell.value][0] > 1:
                # font color = #FF0000
                cell.font = Font(color="FF0000")
                if show_dup_origin:
                    for i, x in enumerate(set(valuesDict[cell.value][1])):
                        ws.cell(
                            row=cell.row,
                            column=cell.column + i + 1,
                            value=x.split(path_delimiter)[-1],
                        )
        if make_copy:
            fp.mkdir(exist_ok=True)
            fp = fp.resolve()
            xp_file_name = xp.split(path_delimiter)[-1]
            wb.save(filename=fp / xp_file_name)
        else:
            wb.save(xp)
        await asyncio.sleep(0)


# ----------------- test ----------------------
async def test() -> None:
    fp = Path(".") / "months"
    xl_files = fp.glob("*.xlsx")
    test_files = set([x.resolve().__str__() for x in xl_files])

    _v = get_files_values(test_files)
    await edit_files_values(
        valuesDict=_v, files=test_files, make_copy=True, show_dup_origin=True, test=True
    )


if __name__ == "__main__":
    generate_days(days=31, numbers=400)
    asyncio.run(test())
